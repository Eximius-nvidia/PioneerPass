"""
This module contains the method that generates summary of test case reports
"""

import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference


def generate_summary(MP_centers=None):
    """
    Method to generate summary
    :param date: for which date the file need to be created ("%Y-%m-%d")
    :return:
    """

    date = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    reports_directory = os.path.join(os.getcwd(), "reports", date)
    file_list = [os.path.join(reports_directory, date + "-" + str(center) + ".log") \
                 for center in MP_centers]
    for file in file_list:
        mp_id = file.split("-")[-1].split(".")[-2]
        file = file.replace(" ", "_")
        try:
            with open(file, "r") as log_file:
                count_pass = 0
                count_fail = 0
                count_total = 0
                for line in log_file.readlines()[10:]:
                    if line.strip()[-4:] == "PASS":
                        count_pass += 1
                    elif line.strip()[-4:] == "FAIL":
                        count_fail += 1
                count_total = count_pass + count_fail
                data_to_append = (date, count_pass, count_fail, count_total)
                # Call the excel summary generator
                excel_summary(mp_id, data_to_append)
        except FileNotFoundError:
            data_to_append = (date, 0, 0, 0)
            # Call the excel summary generator
            excel_summary(mp_id, data_to_append)

def excel_summary(mp_id, data_to_append):
    """
    generate excel summary
    :param mp_id:
    :param data_to_append:
    :return:
    """

    # create summary path
    mp_id = mp_id.replace(" ", "_")
    summary_path = os.path.join(os.getcwd(), "reports", "summary")
    if os.path.isdir(summary_path) == False:
        os.makedirs(summary_path)

    workbook_title = os.path.join(summary_path, "test_summary-{}.xlsx".format(mp_id))

    try:
        wrk_book = load_workbook(workbook_title)
    except FileNotFoundError:
        wrk_book = Workbook()

    if mp_id not in wrk_book.sheetnames:
        wrk_sheet = wrk_book.create_sheet(title=mp_id)
    else:
        wrk_sheet = wrk_book[mp_id]

    if "Sheet" in wrk_book.sheetnames:
        del wrk_book["Sheet"]

    header = ('Date', 'Pass', 'Fail', 'Total',)

    if wrk_sheet.max_row <= 1:
        wrk_sheet.append(header)

    wrk_sheet.append(data_to_append)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Bar Chart"
    chart1.y_axis.title = 'Test number'
    chart1.x_axis.title = 'Date'

    data = Reference(wrk_sheet, min_col=2, min_row=1, max_row=wrk_sheet.max_row, max_col=4)
    cats = Reference(wrk_sheet, min_col=1, min_row=2, max_row=wrk_sheet.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4

    summary_chart = chart1.series[0]
    summary_chart.graphicalProperties.line.solidFill = "00000"
    summary_chart.graphicalProperties.solidFill = "05a308"

    summary_chart = chart1.series[1]
    summary_chart.graphicalProperties.line.solidFill = "00000"
    summary_chart.graphicalProperties.solidFill = "ff9900"

    summary_chart = chart1.series[2]
    summary_chart.graphicalProperties.line.solidFill = "00000"
    summary_chart.graphicalProperties.solidFill = "3853ff"

    wrk_sheet.add_chart(chart1, "G2")
    #print("Summary generated:",workbook_title)
    wrk_book.save(workbook_title)


if __name__ == "__main__":
    yesterday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
    generate_summary(yesterday)
